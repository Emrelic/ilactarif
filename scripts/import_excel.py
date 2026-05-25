"""Eczanenin 7 yıllık ilaç listesini Excel'den SQLite'a yükler.

Çıkış sütununa göre azalan sırada rank atanır (1 = en çok satılan).
Idempotent: tekrar çalıştırıldığında aynı barkodlu kayıt UPDATE edilir.

Kullanım:
    python scripts/import_excel.py
    python scripts/import_excel.py --xlsx "C:/Users/.../İLAÇ TARİF.xls.xlsx"
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

# src/ klasörünü PYTHONPATH'e ekle
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import openpyxl

from ilactarif.db import get_connection, init_schema  # noqa: E402

DEFAULT_XLSX = Path(r"C:\Users\user\OneDrive\Masaüstü\İLAÇ TARİF.xls.xlsx")


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                    help="İlaç listesi Excel dosyası")
    return ap.parse_args()


def read_rows(xlsx_path: Path):
    """Excel satırlarını dict listesi olarak döndürür (başlık satırı atlanır)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        barkod, tip, firma, urun_adi, stok, cikis, toplam, esd = (list(row) + [None] * 8)[:8]
        if not barkod or not urun_adi:
            continue
        rows.append({
            "barcode":         str(barkod).strip(),
            "drug_type":       (tip or "").strip(),
            "manufacturer":    (firma or "").strip(),
            "name":            str(urun_adi).strip(),
            "stock":           int(stok or 0),
            "sales_count":     int(cikis or 0),
            "equivalent_code": str(esd).strip() if esd is not None else "",
        })
    wb.close()
    return rows


def main() -> int:
    args = parse_args()
    if not args.xlsx.exists():
        print(f"HATA: dosya bulunamadı: {args.xlsx}", file=sys.stderr)
        return 1

    print(f"Okunan dosya: {args.xlsx}")
    rows = read_rows(args.xlsx)
    print(f"Okunan satır: {len(rows)}")

    # Çıkış'a göre azalan sırala, rank ata
    rows.sort(key=lambda r: r["sales_count"], reverse=True)
    for i, r in enumerate(rows, start=1):
        r["rank"] = i

    with get_connection() as conn:
        init_schema(conn)
        cur = conn.cursor()
        upserts = 0
        for r in rows:
            cur.execute("""
                INSERT INTO drugs (barcode, name, manufacturer, drug_type,
                                   sales_count, stock, equivalent_code, rank)
                VALUES (:barcode, :name, :manufacturer, :drug_type,
                        :sales_count, :stock, :equivalent_code, :rank)
                ON CONFLICT(barcode) DO UPDATE SET
                    name            = excluded.name,
                    manufacturer    = excluded.manufacturer,
                    drug_type       = excluded.drug_type,
                    sales_count     = excluded.sales_count,
                    stock           = excluded.stock,
                    equivalent_code = excluded.equivalent_code,
                    rank            = excluded.rank,
                    updated_at      = CURRENT_TIMESTAMP
            """, r)
            upserts += 1
        conn.commit()

        total = cur.execute("SELECT COUNT(*) FROM drugs").fetchone()[0]
        active = cur.execute(
            "SELECT COUNT(*) FROM drugs WHERE drug_type = 'İLAÇ'"
        ).fetchone()[0]
        passive = cur.execute(
            "SELECT COUNT(*) FROM drugs WHERE drug_type = 'PASİF İLAÇ'"
        ).fetchone()[0]

    print(f"\nUpsert edilen satır: {upserts}")
    print(f"DB'deki toplam ilaç: {total}  (İLAÇ={active}, PASİF İLAÇ={passive})")
    print("\nEn çok satan 10 ilaç:")
    with get_connection() as conn:
        for r in conn.execute(
            "SELECT rank, name, sales_count FROM drugs ORDER BY rank LIMIT 10"
        ):
            print(f"  #{r['rank']:>4}  {r['sales_count']:>6} adet  {r['name']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
